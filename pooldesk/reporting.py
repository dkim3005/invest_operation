"""Module 7 — reporting.

Turns the day's processed data into the artefacts an operations team actually
circulates:

* a daily Excel ops pack (seven tabs, formatted, with break/trade highlighting)
* a daily one-page PDF summary with a NAV-trend chart
* weekly and monthly roll-ups

Also exposes ``export_for_powerbi`` (used by the Power BI dashboard, Module 11).
See BUILD_SPEC.md Module 7.
"""
from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from pooldesk import data_quality, db, nav, reconcile

# ── styling constants ───────────────────────────────────────────────────────
_TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_LABEL_FONT = Font(bold=True)
_RED_FILL = PatternFill("solid", fgColor="FFC7CE")     # HIGH / FAILED
_YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")  # MEDIUM / PENDING


def _daily_dir() -> Path:
    d = config.REPORTS_DIR / "daily"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _df_sheet(wb: Workbook, name: str, df: pd.DataFrame,
              highlight_col: str | None = None,
              high_values: tuple[str, ...] = (),
              warn_values: tuple[str, ...] = ()) -> None:
    """Write a DataFrame to a new worksheet with a styled header.

    Rows whose ``highlight_col`` value is in ``high_values`` are filled red,
    in ``warn_values`` yellow.
    """
    ws = wb.create_sheet(title=name[:31])
    if df.empty:
        ws["A1"] = f"{name} — no records for this day"
        ws["A1"].font = _LABEL_FONT
        return

    for c, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c, value=str(col))
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = min(
            48, max(12, len(str(col)) + 2,
                    int(df[col].astype(str).str.len().max()) + 2))

    hl_idx = list(df.columns).index(highlight_col) if highlight_col else None
    for r, (_, row) in enumerate(df.iterrows(), start=2):
        fill = None
        if hl_idx is not None:
            value = str(row.iloc[hl_idx])
            if value in high_values:
                fill = _RED_FILL
            elif value in warn_values:
                fill = _YELLOW_FILL
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c,
                           value=value.item() if hasattr(value, "item")
                           else value)
            if fill:
                cell.fill = fill
    ws.freeze_panes = "A2"


def _summary_sheet(ws, run_date: date) -> None:
    """Key-value overview tab."""
    diso = run_date.isoformat()
    score = data_quality.dq_scorecard(run_date)["overall_score"]
    total_nav = float(db.query(
        "SELECT COALESCE(SUM(nav_cad),0) AS s FROM fact_nav WHERE nav_date=?",
        (diso,)).s.iloc[0])
    sev = dict(db.query(
        "SELECT severity, COUNT(*) AS n FROM recon_exception WHERE run_date=? "
        "GROUP BY severity", (diso,)).itertuples(index=False, name=None))
    trades = dict(db.query(
        "SELECT status, COUNT(*) AS n FROM fact_trade WHERE feed_date=? "
        "GROUP BY status", (diso,)).itertuples(index=False, name=None))

    ws["A1"] = "PoolDesk — Daily Operations Pack"
    ws["A1"].font = _TITLE_FONT
    rows = [
        ("Business date", diso),
        ("Data quality score (0-100)", score),
        ("Total NAV (CAD)", round(total_nav, 2)),
        ("Reconciliation breaks — HIGH", sev.get("HIGH", 0)),
        ("Reconciliation breaks — MEDIUM", sev.get("MEDIUM", 0)),
        ("Reconciliation breaks — LOW", sev.get("LOW", 0)),
        ("Trades — total", sum(trades.values())),
        ("Trades — FAILED", trades.get("FAILED", 0)),
        ("Trades — PENDING", trades.get("PENDING", 0)),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24


def build_daily_excel(run_date: date) -> Path:
    """Build the six-tab daily operations workbook."""
    diso = run_date.isoformat()
    wb = Workbook()
    _summary_sheet(wb.active, run_date)
    wb.active.title = "Summary"

    _df_sheet(wb, "DQ Scorecard", db.query(
        "SELECT check_name, severity, records_checked, records_failed, "
        "pass_rate, detail FROM dq_result WHERE run_date=?", (diso,)))

    _df_sheet(wb, "Reconciliation Exceptions", db.query(
        "SELECT break_id, pool_id, security_id, break_type, qty_diff, "
        "mv_impact_cad, severity, ai_priority, ai_root_cause, ai_owner_team, "
        "status, ai_resolution_note FROM recon_exception WHERE run_date=? "
        "ORDER BY mv_impact_cad DESC", (diso,)),
        highlight_col="severity", high_values=("HIGH",),
        warn_values=("MEDIUM",))

    navdf = db.query(
        "SELECT n.pool_id, p.pool_name, n.gross_asset_value_cad, "
        "n.liabilities_cad, n.nav_cad, n.units_outstanding, n.unit_price "
        "FROM fact_nav n JOIN dim_pool p ON p.pool_id=n.pool_id "
        "WHERE n.nav_date=?", (diso,))
    if not navdf.empty:
        navdf = navdf.merge(nav.daily_pnl(run_date)[["pool_id",
                            "daily_pnl_cad"]], on="pool_id", how="left")
    _df_sheet(wb, "NAV by Pool", navdf)

    _df_sheet(wb, "Cash Reconciliation", reconcile.reconcile_cash(run_date),
              highlight_col="status", high_values=("BREAK",))

    _df_sheet(wb, "Client Holdings", db.query(
        "SELECT h.client_id, c.client_name, h.pool_id, h.units_held, "
        "h.holding_value_cad FROM fact_client_holding h "
        "JOIN dim_client c ON c.client_id=h.client_id "
        "WHERE h.holding_date=? ORDER BY h.holding_value_cad DESC", (diso,)))

    _df_sheet(wb, "Trade Blotter", db.query(
        "SELECT trade_id, trade_date, pool_id, security_id, side, quantity, "
        "price, settlement_date, status FROM fact_trade WHERE feed_date=? "
        "ORDER BY trade_id", (diso,)),
        highlight_col="status", high_values=("FAILED",),
        warn_values=("PENDING",))

    out = _daily_dir() / f"PoolDesk_Ops_{run_date:%Y%m%d}.xlsx"
    wb.save(out)
    return out


def _nav_trend_chart(run_date: date) -> Path:
    """Render a total-NAV trend line up to run_date; return the PNG path."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    trend = db.query(
        "SELECT nav_date, SUM(nav_cad) AS total FROM fact_nav "
        "WHERE nav_date<=? GROUP BY nav_date ORDER BY nav_date",
        (run_date.isoformat(),))
    fig, ax = plt.subplots(figsize=(7.2, 2.8))
    ax.plot(trend.nav_date, trend.total / 1e6, marker="o", color="#1F4E78")
    ax.set_title("Total NAV trend (CAD millions)")
    ax.set_ylabel("CAD m")
    ax.tick_params(axis="x", rotation=45, labelsize=7)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    png = Path(tempfile.gettempdir()) / f"pooldesk_navtrend_{run_date:%Y%m%d}.png"
    fig.savefig(png, dpi=120)
    plt.close(fig)
    return png


def build_daily_pdf(run_date: date) -> Path:
    """Build the one-page daily PDF summary."""
    from fpdf import FPDF

    diso = run_date.isoformat()
    score = data_quality.dq_scorecard(run_date)["overall_score"]
    total_nav = float(db.query(
        "SELECT COALESCE(SUM(nav_cad),0) AS s FROM fact_nav WHERE nav_date=?",
        (diso,)).s.iloc[0])
    n_breaks = int(db.query(
        "SELECT COUNT(*) AS n FROM recon_exception WHERE run_date=?",
        (diso,)).n.iloc[0])
    top = db.query(
        "SELECT pool_id, security_id, break_type, mv_impact_cad, severity, "
        "ai_root_cause, ai_owner_team FROM recon_exception WHERE run_date=? "
        "ORDER BY mv_impact_cad DESC LIMIT 5", (diso,))

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "PoolDesk - Daily Operations Summary",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 7, f"Business date: {diso}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Key indicators", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    for label, value in [
        ("Data quality score", f"{score} / 100"),
        ("Total NAV (CAD)", f"{total_nav:,.0f}"),
        ("Reconciliation breaks", str(n_breaks)),
    ]:
        pdf.cell(60, 7, label)
        pdf.cell(0, 7, value, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    chart = _nav_trend_chart(run_date)
    pdf.image(str(chart), w=180)
    chart.unlink(missing_ok=True)
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Top breaks by market-value impact",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    if top.empty:
        pdf.cell(0, 6, "No reconciliation breaks.",
                 new_x="LMARGIN", new_y="NEXT")
    else:
        for t in top.itertuples(index=False):
            line = (f"{t.severity:<7} {t.pool_id:<11} {t.security_id:<8} "
                    f"{t.break_type:<22} CAD {t.mv_impact_cad:>13,.0f}  "
                    f"-> {t.ai_root_cause} / {t.ai_owner_team}")
            pdf.cell(0, 6, line, new_x="LMARGIN", new_y="NEXT")

    out = _daily_dir() / f"PoolDesk_Summary_{run_date:%Y%m%d}.pdf"
    pdf.output(str(out))
    return out


def _period_metrics(days: list[date]) -> pd.DataFrame:
    """Per-day headline metrics across a set of business days."""
    rows = []
    for d in days:
        diso = d.isoformat()
        if db.query("SELECT COUNT(*) AS n FROM fact_nav WHERE nav_date=?",
                    (diso,)).n.iloc[0] == 0:
            continue
        sev = dict(db.query(
            "SELECT severity, COUNT(*) AS n FROM recon_exception "
            "WHERE run_date=? GROUP BY severity",
            (diso,)).itertuples(index=False, name=None))
        rows.append({
            "date": diso,
            "dq_score": data_quality.dq_scorecard(d)["overall_score"],
            "total_nav_cad": round(float(db.query(
                "SELECT SUM(nav_cad) AS s FROM fact_nav WHERE nav_date=?",
                (diso,)).s.iloc[0]), 2),
            "breaks_total": sum(sev.values()),
            "breaks_high": sev.get("HIGH", 0),
        })
    return pd.DataFrame(rows)


def _rollup(period: str, days: list[date], label: str) -> Path:
    """Shared weekly/monthly roll-up writer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Trend"
    ws["A1"] = f"PoolDesk - {period} roll-up - {label}"
    ws["A1"].font = _TITLE_FONT
    metrics = _period_metrics(days)
    if metrics.empty:
        ws["A3"] = "No processed data in this period."
    else:
        for c, col in enumerate(metrics.columns, start=1):
            cell = ws.cell(row=3, column=c, value=col)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            ws.column_dimensions[get_column_letter(c)].width = 18
        for r, (_, row) in enumerate(metrics.iterrows(), start=4):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)
        avg_row = 4 + len(metrics) + 1
        ws.cell(row=avg_row, column=1, value="Average DQ score").font = _LABEL_FONT
        ws.cell(row=avg_row, column=2,
                value=round(float(metrics.dq_score.mean()), 1))
    out = config.REPORTS_DIR / period.lower()
    out.mkdir(parents=True, exist_ok=True)
    path = out / f"PoolDesk_{period}_{label}.xlsx"
    wb.save(path)
    return path


def build_weekly_rollup(week_end: date) -> Path:
    """Roll up the (up to) five business days ending on week_end."""
    days = [d for d in config.business_days(config.SIM_START_DATE,
            config.SIM_DAYS) if d <= week_end][-5:]
    return _rollup("Weekly", days, week_end.strftime("%Y%m%d"))


def build_monthly_rollup(month: date) -> Path:
    """Roll up every business day in the calendar month of ``month``."""
    days = [d for d in config.business_days(config.SIM_START_DATE,
            config.SIM_DAYS)
            if d.year == month.year and d.month == month.month]
    return _rollup("Monthly", days, month.strftime("%Y%m"))


def export_for_powerbi() -> Path:
    """Dump the analytical tables to CSV for the Power BI dashboard."""
    out = config.REPORTS_DIR / "powerbi"
    out.mkdir(parents=True, exist_ok=True)
    exports = {
        "fact_nav": "SELECT * FROM fact_nav",
        "fact_client_holding": "SELECT * FROM fact_client_holding",
        "recon_exception": "SELECT * FROM recon_exception",
        "dq_result": "SELECT * FROM dq_result",
        "fact_trade": "SELECT * FROM fact_trade",
        # Dimensions so a CSV-only Power BI model can resolve names/types
        # and filter every fact by a shared date.
        "dim_pool": "SELECT * FROM dim_pool",
        "dim_client": "SELECT * FROM dim_client",
        "dim_security": "SELECT * FROM dim_security",
        "dim_date": "SELECT * FROM dim_date",
    }
    for name, query in exports.items():
        db.query(query).to_csv(out / f"{name}.csv", index=False)
    print(f"[reporting] exported {len(exports)} tables for Power BI to {out}")
    return out
